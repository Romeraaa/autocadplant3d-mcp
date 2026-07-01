"""Volcado de resultados de extraccion a CSV / XLSX.

Columnas del line-list = union ordenada de los campos de :class:`~autocad_mcp.pnid.extract.LineRecord`.
Ademas genera un informe de cobertura (``COBERTURA.txt``) y, en XLSX, una hoja de tokens
no reconocidos para revision del ingeniero.
"""

from __future__ import annotations

import csv
import os

from .extract import SheetResult

# Orden estable de columnas del line-list.
LINE_COLUMNS = [
    "sheet",
    "line_id",
    "family",
    "diameter",
    "service",
    "area",
    "number",
    "clase",
    "name",
    "page",
    "x",
    "y",
    "count",
]


def _line_rows(results: list[SheetResult]):
    """Genera filas (dict) del line-list ordenadas por sheet y line_id."""
    rows = []
    for res in results:
        for rec in res.lines:
            rows.append({c: getattr(rec, c) for c in LINE_COLUMNS})
    rows.sort(key=lambda r: (r["sheet"], r["line_id"]))
    return rows


def coverage_report(results: list[SheetResult]) -> str:
    """Texto del informe de cobertura por hoja + total."""
    lines_out = ["INFORME DE COBERTURA - line-list desde P&ID PDF", "=" * 50, ""]
    tot_cand = tot_reco = tot_lineas = 0
    for res in results:
        tot_cand += res.candidates
        tot_reco += res.recognized
        tot_lineas += len(res.lines)
        lines_out.append(
            f"[{res.sheet}] lineas unicas: {len(res.lines):3d} | "
            f"candidatos: {res.candidates:4d} | reconocidos: {res.recognized:4d} | "
            f"cobertura: {res.coverage * 100:5.1f}%"
        )
        if res.unrecognized:
            lines_out.append(f"    no reconocidos ({len(res.unrecognized)}):")
            for tok in res.unrecognized:
                lines_out.append(f"      - {tok}")
        lines_out.append("")
    total_cov = 1.0 if tot_cand == 0 else tot_reco / tot_cand
    lines_out.append("-" * 50)
    lines_out.append(
        f"TOTAL  lineas unicas: {tot_lineas} | candidatos: {tot_cand} | "
        f"reconocidos: {tot_reco} | cobertura global: {total_cov * 100:.1f}%"
    )
    return "\n".join(lines_out)


def write_csv(results: list[SheetResult], out_dir: str) -> list[str]:
    """Escribe ``LINE_LIST.csv`` + ``COBERTURA.txt`` (+ bonus si hay). Devuelve rutas."""
    os.makedirs(out_dir, exist_ok=True)
    written = []

    line_path = os.path.join(out_dir, "LINE_LIST.csv")
    with open(line_path, "w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=LINE_COLUMNS)
        writer.writeheader()
        writer.writerows(_line_rows(results))
    written.append(line_path)

    written.append(write_coverage_txt(results, out_dir))
    return written


def write_coverage_txt(results: list[SheetResult], out_dir: str) -> str:
    """Escribe ``COBERTURA.txt`` en ``out_dir`` y devuelve su ruta."""
    os.makedirs(out_dir, exist_ok=True)
    cov_path = os.path.join(out_dir, "COBERTURA.txt")
    with open(cov_path, "w", encoding="utf-8") as fh:
        fh.write(coverage_report(results))
    return cov_path


def write_xlsx(results: list[SheetResult], out_dir: str) -> list[str]:
    """Escribe ``LINE_LIST.xlsx`` (hojas: Lineas, NoReconocidos, Cobertura). Devuelve rutas."""
    from openpyxl import Workbook

    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, "LINE_LIST.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Lineas"
    ws.append(LINE_COLUMNS)
    for row in _line_rows(results):
        ws.append([row[c] for c in LINE_COLUMNS])

    ws_nr = wb.create_sheet("NoReconocidos")
    ws_nr.append(["sheet", "token"])
    for res in results:
        for tok in res.unrecognized:
            ws_nr.append([res.sheet, tok])

    ws_cov = wb.create_sheet("Cobertura")
    ws_cov.append(["sheet", "lineas_unicas", "candidatos", "reconocidos", "cobertura_%"])
    for res in results:
        ws_cov.append(
            [res.sheet, len(res.lines), res.candidates, res.recognized, round(res.coverage * 100, 1)]
        )

    # Bonus opcional: solo si hay datos.
    if any(res.instruments for res in results):
        ws_i = wb.create_sheet("Instrumentos")
        ws_i.append(["sheet", "page", "tag", "prefix", "func", "number"])
        for res in results:
            for it in res.instruments:
                ws_i.append([it["sheet"], it["page"], it["tag"], it["prefix"], it["func"], it["number"]])
    if any(res.equipment for res in results):
        ws_e = wb.create_sheet("Equipos")
        ws_e.append(["sheet", "page", "tag", "a", "b", "c"])
        for res in results:
            for eq in res.equipment:
                ws_e.append([eq["sheet"], eq["page"], eq["tag"], eq["a"], eq["b"], eq["c"]])

    wb.save(path)
    return [path]
