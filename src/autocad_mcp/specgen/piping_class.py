"""Parse a REPSOL piping-class Excel into a flat list of normalised entries.

Standalone proof of concept (NOT part of the MCP server). Standard library plus ``openpyxl``.

This module ONLY parses (no catalog access, no matching). It discovers the family sheets and the
column layout from the headers, so it is not tied to one particular workbook. Each parsed row
becomes a :class:`PipingClassEntry`.

Key facts driving the parser (verified against the sample piping class):
  * Every family sheet carries an explicit ``L CODE`` column. That column is the canonical source
    of the REPSOL L/H code. Some sheets (e.g. ``STUD-BOLT``) do NOT repeat the code inside the
    bilingual ``DESCRIPCION`` text, so reading it from the description alone misses them
    (this was the STUD-BOLT / H-291..H-298 bug). We therefore read the dedicated column first and
    fall back to the description only when the column is absent/empty.
  * The hydrogen variant is encoded as an ``-H2`` suffix on the L-code (``L-1276-H2``); the base
    code is the same token without the suffix. This is detected generically (see
    ``common.is_variant_code`` / ``common.base_lcode``), not from a hard-coded list.
"""

from __future__ import annotations

from dataclasses import dataclass, field

from openpyxl import load_workbook

from . import common
from .common import norm

# Sheets that are NOT component families (cover / notes / limits). Everything else is parsed.
NON_FAMILY_SHEETS = {"NXD-2", "NOTAS DE REFERENCIA"}

# End-type wording (bilingual) -> catalog EndType vocab (WF, PL, THDF, SW, BV, THDM, FL, GRV).
# Checked in priority order against a normalised (accent-free, lower-case) description string.
END_PATTERNS: list[tuple[str, str]] = [
    ("embocadura para soldar", "SW"),     # socket weld
    ("socket weld", "SW"),
    ("biselad", "BV"),                     # bevelled (butt weld bevel)
    ("butt weld", "BV"),
    ("soldadura a tope", "BV"),
    ("extremos planos", "PL"),             # plain ends
    ("plain end", "PL"),
    ("extremo plano", "PL"),
    ("roscad", "THDF"),                    # threaded (female by default)
    ("threaded", "THDF"),
    ("bridad", "FL"),                      # flanged
    ("flanged", "FL"),
    ("cara con resalte", "FL"),            # raised face flange
    ("entre bridas", "WF"),                # wafer / between flanges (gaskets, blinds)
    ("cuello para soldar", "FL"),          # welding neck flange
]


@dataclass
class PipingClassEntry:
    """One row of the piping class Excel, normalised."""

    sheet: str
    family: str                 # column FAMILIA
    type_: str                  # column TIPO
    unicode_code: str           # column UNICODE (REPSOL part code)
    description: str            # column DESCRIPCION (bilingual ES/EN)
    lcode: str | None           # REPSOL L-/H- code (full, including any -H2 suffix)
    lcode_base: str | None      # L-code with the variant (-H2) suffix stripped
    is_hydrogen: bool           # the L-code carries the variant suffix
    main_diameter: float | None    # Ø MA. (inches)
    branch_diameter: float | None  # Ø ME. (inches), for olets/swages/reducers
    schedule: str | None        # SCH MA. (normalised)
    rating: str | None          # RATING / PressureClass (normalised, digits only)
    end_type: str | None        # deduced: PL / BV / SW / THDF / FL / WF ...

    # matching outcome (filled by the matcher)
    catalog: str | None = None
    catalog_class: str | None = None
    family_desc: str | None = None
    size_record_id: bytes | None = None
    part_family_id: bytes | None = None
    pnpid: int | None = None
    match_note: str = ""

    # confidence outcome (filled by the matcher's scorer)
    confidence: str = "BAJA"            # ALTA | MEDIA | SUSTITUCION | BAJA
    score: float = 0.0
    alternatives: list[tuple[str, float, str | None, str | None]] = field(default_factory=list)


def _deduce_end_type(description: str) -> str | None:
    n = norm(description)
    for pattern, end in END_PATTERNS:
        if pattern in n:
            return end
    return None


def _header_index(header_row: tuple) -> tuple[dict[str, int], list[str]]:
    """Map a normalised header label to its column index; also return duplicate labels (H7).

    The FIRST occurrence of each label wins (``setdefault``); any further occurrence of the same
    normalised label is reported back so the caller can warn (a duplicated header silently shadows
    a column otherwise).
    """
    idx: dict[str, int] = {}
    duplicates: list[str] = []
    for i, cell in enumerate(header_row):
        key = norm(cell)
        if not key:
            continue
        if key in idx:
            duplicates.append(key)
        else:
            idx[key] = i
    return idx, duplicates


def _pick(idx: dict[str, int], *aliases: str) -> int | None:
    """Resolve a column index by exact normalised label, then by substring fallback."""
    for a in aliases:
        if a in idx:
            return idx[a]
    for key, col in idx.items():
        for a in aliases:
            if a and a in key:
                return col
    return None


def _cell(row: tuple, col: int | None) -> object | None:
    if col is None or col >= len(row):
        return None
    return row[col]


def _cell_str(row: tuple, col: int | None) -> str:
    v = _cell(row, col)
    return str(v).strip() if v is not None and str(v).strip() else ""


def parse_workbook(path: str, *, warn=None) -> list[PipingClassEntry]:
    """Parse every component-family sheet into a flat list of :class:`PipingClassEntry`.

    ``warn`` is an optional ``callable(str)`` used for non-fatal notices (H7 duplicate headers).
    Sheets in :data:`NON_FAMILY_SHEETS` and sheets without a ``descripcion`` header are skipped.
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    entries: list[PipingClassEntry] = []
    try:
        for sheet in wb.sheetnames:
            if sheet in NON_FAMILY_SHEETS:
                continue
            ws = wb[sheet]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            header_i = next(
                (i for i, r in enumerate(rows) if any("descripcion" in norm(c) for c in r)),
                None,
            )
            if header_i is None:
                continue   # not a component-family sheet
            idx, dups = _header_index(rows[header_i])
            if dups and warn is not None:
                warn(f"hoja '{sheet}': encabezados duplicados {sorted(set(dups))} "
                     f"(se usa la primera aparicion)")

            c_fam = _pick(idx, "familia")
            c_type = _pick(idx, "tipo")
            c_uni = _pick(idx, "unicode")
            c_desc = _pick(idx, "descripcion")
            c_dma = _pick(idx, "ma.", "ma", "diam ma")       # Ø MA.
            c_dme = _pick(idx, "me.", "me", "diam me")        # Ø ME.
            c_sch = _pick(idx, "sch ma.", "sch ma", "sch")
            c_rating = _pick(idx, "rating", "sch ma./rating")
            c_lcode = _pick(idx, "l code", "l-code", "lcode", "codigo l", "code")
            if c_desc is None:
                continue

            for r in rows[header_i + 1:]:
                desc = _cell(r, c_desc)
                if not desc or not str(desc).strip():
                    continue
                description = str(desc)

                # L-code: dedicated column first (canonical), then the description text.
                lcode = common.extract_lcode(_cell(r, c_lcode))
                if not lcode:
                    lcode = common.extract_lcode(description)
                is_var = common.is_variant_code(lcode)
                lcode_base = common.base_lcode(lcode)

                entries.append(PipingClassEntry(
                    sheet=sheet,
                    family=_cell_str(r, c_fam),
                    type_=_cell_str(r, c_type),
                    unicode_code=_cell_str(r, c_uni),
                    description=description,
                    lcode=lcode,
                    lcode_base=lcode_base,
                    is_hydrogen=is_var,
                    main_diameter=common.parse_diameter(_cell(r, c_dma)),
                    branch_diameter=common.parse_diameter(_cell(r, c_dme)),
                    schedule=common.norm_schedule(_cell(r, c_sch)),
                    rating=common.norm_rating(_cell(r, c_rating)),
                    end_type=_deduce_end_type(description),
                ))
    finally:
        wb.close()
    return entries
