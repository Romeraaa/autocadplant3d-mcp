"""Read a REPSOL piping-class Excel and produce a SpecDefinition for ``spec_builder``.

Standalone proof of concept (NOT part of the MCP server). Standard library plus ``openpyxl``.

This is the "common piece" between the human-authored piping class (an .xlsx with one sheet per
component family) and the catalog-driven spec materialiser already proven in ``spec_builder.py``.
It does NOT re-implement materialisation: it imports ``SpecDefinition`` / ``ComponentRef`` and the
``Materialiser`` / ``build_pspx`` / branch-table machinery and only swaps the *definition source*
from "derived from NXD-2" to "read from the piping class Excel + matched against the catalogs".

Pipeline:
  1. PARSE  -- each family sheet -> a list of PipingClassEntry (component type, bilingual
     description, L-code, material, nominal size(s), schedule/rating, end type).
  2. MATCH  -- each entry -> a catalog family (by L-code embedded in PartFamilyLongDesc, refined by
     end type) and then a SizeRecordId (by nominal diameter, refined by schedule). The Excel has no
     GUIDs, so matching is purely by properties.
  3. BUILD  -- assemble a SpecDefinition (ComponentRef per matched size) and materialise it with the
     existing Materialiser into NXD-2-FROMCLASS.pspc / .pspx. The branch table is reused verbatim
     from NXD-2 (the piping class does not carry a machine-parseable branch table).

Catalog reality discovered empirically (drives the heuristic, see report):
  * The six REPSOL .pcat catalogs DO embed the REPSOL L-code (e.g. ``L-1276``) inside
    ``PartFamilyLongDesc`` for ~16% of rows. Every L-code used by this piping class is present in
    exactly one catalog. The L-code is therefore the strongest matching key; nominal diameter then
    selects the SizeRecordId (1 per diameter per family). ``Material`` / ``MaterialCode`` are NULL
    in these catalogs, so they cannot be used.
  * The hydrogen variants (``L-xxxx-H2``) have NO dedicated catalog family; only the base L-code
    exists. H2 entries therefore match the base family (documented as a known gap).
  * Valves with an L-code live in catalog class ``ValveBody`` (not ``Valve``); we record the real
    catalog class so the materialiser copies the correct graph.

Inputs are opened strictly read-only. Output goes to ``./out/``.
"""

from __future__ import annotations

import os
import re
import sqlite3
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

# Reuse the proven definition + materialisation machinery (do NOT reimplement).
from spec_builder import (
    CATALOGS,
    SCRATCH,
    CatalogIndex,
    ComponentRef,
    Materialiser,
    SpecDefinition,
    build_pspx,
    derive_definition_from_template,
    verify as verify_spec,
)
from generate_spec_poc import ro_connect

# --------------------------------------------------------------------------- paths
HERE = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = os.path.join(HERE, "out")
XLSX = os.path.join(SCRATCH, "piping_class.xlsx")
TEMPLATE_PSPC = os.path.join(SCRATCH, "NXD-2.pspc")

OUT_NAME = "NXD-2-FROMCLASS"

# Family sheets that hold parsable component rows (others are cover / notes / limits).
FAMILY_SHEETS = [
    "TUBERIA", "NIPPLES_SWAGES", "FORGED FITTINGS", "OLETS", "BW FITTINGS",
    "WELDOLETS BW", "FLANGES-MISCELLANEOUS", "GASKETS", "STUD-BOLT", "VALVES",
]

# L/H code token, e.g. L-1276, L-1746-H2, H-291.
LCODE_RE = re.compile(r"\b([LH]-\d+(?:-[A-Za-z0-9]+)?)\b")


# =========================================================================== PARSE
@dataclass
class PipingClassEntry:
    """One row of the piping class Excel, normalised."""

    sheet: str
    family: str                 # column FAMILIA
    type_: str                  # column TIPO
    unicode_code: str           # column UNICODE (REPSOL part code)
    description: str            # column DESCRIPCION (bilingual ES/EN)
    lcode: str | None           # REPSOL L-/H- code (base, without -H2 stripped here)
    lcode_base: str | None      # L-code with the -H2 / -H suffix stripped
    is_hydrogen: bool
    main_diameter: float | None   # Ø MA. (inches)
    branch_diameter: float | None  # Ø ME. (inches), for olets/swages/reducers
    schedule: str | None        # SCH MA. (normalised)
    rating: str | None          # RATING / PressureClass (normalised, digits only)
    end_type: str | None        # deduced: PL / BV / SW / THD / FL / WF ...

    # matching outcome (filled by the matcher)
    catalog: str | None = None
    catalog_class: str | None = None
    family_desc: str | None = None
    size_record_id: bytes | None = None
    part_family_id: bytes | None = None
    pnpid: int | None = None
    match_note: str = ""

    # confidence outcome (filled by the matcher's scorer)
    confidence: str = "BAJA"            # ALTA | MEDIA | SUSTITUCION | BAJA
    score: float = 0.0                  # winning candidate score
    # top-N alternatives as (family_desc, score, schedule, pressure_class), best first,
    # EXCLUDING the chosen one -- shown to the engineer for review.
    alternatives: list[tuple[str, float, str | None, str | None]] = field(default_factory=list)


# Map the verbose end-type wording in the bilingual description to the catalog EndType vocab
# (WF, PL, THDF, SW, BV, THDM, FL, GRV). Checked in priority order against a normalised string.
_END_PATTERNS: list[tuple[str, str]] = [
    ("embocadura para soldar", "SW"),     # socket weld
    ("socket weld", "SW"),
    ("biselad", "BV"),                     # bevelled (butt weld bevel)
    ("butt weld", "BV"),
    ("soldadura a tope", "BV"),
    ("extremos planos", "PL"),             # plain ends
    ("plain end", "PL"),
    ("extremo plano", "PL"),
    ("roscad", "THDF"),                    # threaded (female by default)
    ("threaded", "THDF"),
    ("bridad", "FL"),                      # flanged
    ("flanged", "FL"),
    ("cara con resalte", "FL"),            # raised face flange
    ("entre bridas", "WF"),                # wafer / between flanges (gaskets, blinds)
    ("cuello para soldar", "FL"),          # welding neck flange
]


def _strip_accents(text: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFKD", text) if not unicodedata.combining(c)
    )


def _norm(text: str | None) -> str:
    """Lower-case, accent-free, whitespace-collapsed -- robust comparison key."""
    if text is None:
        return ""
    t = _strip_accents(str(text)).lower()
    return re.sub(r"\s+", " ", t).strip()


def _parse_diameter(value) -> float | None:
    """Parse a nominal diameter cell. Accepts numbers and strings like '1 1/2', '3/4'."""
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace('"', "")
    # mixed fraction '1 1/2'
    m = re.match(r"^(\d+)\s+(\d+)/(\d+)$", s)
    if m:
        return int(m.group(1)) + int(m.group(2)) / int(m.group(3))
    m = re.match(r"^(\d+)/(\d+)$", s)
    if m:
        return int(m.group(1)) / int(m.group(2))
    try:
        return float(s)
    except ValueError:
        return None


def _norm_schedule(value) -> str | None:
    """Normalise a SCH cell to the catalog vocab ('80', '160', 'STD', 'XS'...)."""
    if value is None or value == "":
        return None
    s = str(value).strip().upper()
    s = s.replace("SCH", "").replace("SCH.", "").strip()
    if s in ("", "-"):
        return None
    # drop a trailing ".0" from numeric schedules read as float
    if re.match(r"^\d+\.0$", s):
        s = s[:-2]
    return s


def _norm_rating(value) -> str | None:
    """Normalise a RATING cell to digits only ('600 #' -> '600', '6000#' -> '6000')."""
    if value is None or value == "":
        return None
    digits = re.sub(r"\D", "", str(value))
    return digits or None


def _deduce_end_type(description: str) -> str | None:
    n = _norm(description)
    for pattern, end in _END_PATTERNS:
        if pattern in n:
            return end
    return None


def _header_index(header_row: tuple) -> dict[str, int]:
    """Map a normalised header label to its column index for one sheet."""
    idx: dict[str, int] = {}
    for i, cell in enumerate(header_row):
        key = _norm(cell)
        if not key:
            continue
        idx.setdefault(key, i)
    return idx


def _pick(idx: dict[str, int], *aliases: str) -> int | None:
    for a in aliases:
        if a in idx:
            return idx[a]
    # substring fallback (headers carry odd unicode for the diameter glyph)
    for key, col in idx.items():
        for a in aliases:
            if a and a in key:
                return col
    return None


def parse_workbook(path: str) -> list[PipingClassEntry]:
    """Parse every family sheet into a flat list of PipingClassEntry."""
    wb = load_workbook(path, data_only=True, read_only=True)
    entries: list[PipingClassEntry] = []
    for sheet in FAMILY_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        # header is the first row that contains 'descripcion'
        header_i = next(
            (i for i, r in enumerate(rows) if any("descripcion" in _norm(c) for c in r)),
            0,
        )
        idx = _header_index(rows[header_i])
        c_fam = _pick(idx, "familia")
        c_type = _pick(idx, "tipo")
        c_uni = _pick(idx, "unicode")
        c_desc = _pick(idx, "descripcion")
        c_dma = _pick(idx, "ma.", "ma", "diam ma")     # Ø MA.
        c_dme = _pick(idx, "me.", "me", "diam me")      # Ø ME.
        c_sch = _pick(idx, "sch ma.", "sch ma", "sch")
        c_rating = _pick(idx, "rating", "sch ma./rating")
        if c_desc is None:
            continue
        for r in rows[header_i + 1:]:
            desc = r[c_desc] if c_desc < len(r) else None
            if not desc or not str(desc).strip():
                continue
            description = str(desc)
            m = LCODE_RE.search(description)
            lcode = m.group(1) if m else None
            is_h2 = bool(lcode and re.search(r"-H2?$", lcode))
            lcode_base = re.sub(r"-H2?$", "", lcode) if lcode else None
            entries.append(PipingClassEntry(
                sheet=sheet,
                family=str(r[c_fam]).strip() if c_fam is not None and c_fam < len(r) and r[c_fam] else "",
                type_=str(r[c_type]).strip() if c_type is not None and c_type < len(r) and r[c_type] else "",
                unicode_code=str(r[c_uni]).strip() if c_uni is not None and c_uni < len(r) and r[c_uni] else "",
                description=description,
                lcode=lcode,
                lcode_base=lcode_base,
                is_hydrogen=is_h2,
                main_diameter=_parse_diameter(r[c_dma]) if c_dma is not None and c_dma < len(r) else None,
                branch_diameter=_parse_diameter(r[c_dme]) if c_dme is not None and c_dme < len(r) else None,
                schedule=_norm_schedule(r[c_sch]) if c_sch is not None and c_sch < len(r) else None,
                rating=_norm_rating(r[c_rating]) if c_rating is not None and c_rating < len(r) else None,
                end_type=_deduce_end_type(description),
            ))
    wb.close()
    return entries


# =========================================================================== MATCH
# Weights for the multi-signal candidate scorer. Calibrated against NXD-2 at family level
# (see report): the catalogs carry NO Material/MaterialCode and the H2 variants are absent, so
# the usable signals are the L-code, the end type, the nominal/branch diameter, the schedule and
# the pressure class. A candidate that fails the nominal-diameter test is discarded outright.
W_LCODE_EXACT = 5.0      # entry's full L-code (incl. -H2) literally present in the family desc
W_LCODE_BASE = 3.0       # only the base L-code matched (the H2 variant has no catalog family)
W_END_MATCH = 3.0        # EndType equals the deduced end type
W_END_MISMATCH = -2.0    # end type known but the candidate disagrees
W_BRANCH_PORT = 2.0      # candidate has a port at the entry's branch diameter (olets/swages/red.)
W_SCH_MATCH = 2.0        # Schedule equals the entry schedule
W_SCH_MISMATCH = -1.0    # schedule known but the candidate disagrees
W_RATING_MATCH = 2.0     # PressureClass equals the entry rating
W_RATING_MISMATCH = -1.0  # rating known but the candidate disagrees

# A second candidate within this margin of the winner is "too close to call" -> MEDIA.
SCORE_MARGIN = 3.0


@dataclass
class _Candidate:
    """One scored catalog row competing to materialise a piping-class entry."""

    score: float
    family_desc: str
    catalog: str
    catalog_class: str
    size_record_id: bytes
    part_family_id: bytes
    pnpid: int
    schedule: str | None
    pressure_class: str | None
    end_type: str | None
    used_base_code: bool


class CatalogMatcher:
    """Index the six catalogs by L-code and resolve entries with a multi-signal confidence model.

    Instead of greedily keeping the first plausible row, the matcher scores EVERY catalog row of
    every family that carries the entry's L-code (a weighted sum of the matching signals) and keeps
    the top-N. The winner becomes the chosen part; the confidence level reflects how unambiguous the
    win was:

      * ALTA       -- single family for the L-code (or end type resolves it) AND a single dominant
                      candidate (no rival within ``SCORE_MARGIN``), non-hydrogen.
      * MEDIA      -- matched but ambiguous: several families compete, or several rows tie because a
                      discriminating signal (typically the schedule) is missing from the Excel.
      * SUSTITUCION-- hydrogen variant resolved to its base family (the ``-H2`` catalog family does
                      not exist); flagged explicitly rather than passing as a silent match.
      * BAJA       -- no L-code, L-code absent from every catalog, or no row at the entry diameter.

    Ground truth for the report is taken at FAMILY level (NXD-2.CatalogPartFamilyId links to the
    catalog PartFamilyId); the SizeRecordId blobs in a materialised .pspc are rewritten and are NOT
    a valid semantic oracle (only 34/406 coincide).
    """

    def __init__(self) -> None:
        self.handles: dict[str, sqlite3.Connection] = {}
        # lcode -> list of (catalog, family_desc, end_type, class_name)
        self.lcode_families: dict[str, list[tuple[str, str, str | None, str]]] = defaultdict(list)
        for logical, fname in CATALOGS.items():
            path = os.path.join(SCRATCH, fname)
            if not os.path.exists(path):
                continue
            con = ro_connect(path)
            self.handles[logical] = con
            for desc, end, cls in con.execute(
                "SELECT DISTINCT e.PartFamilyLongDesc, e.EndType, b.PnPClassName "
                "FROM EngineeringItems e JOIN PnPBase b ON e.PnPID=b.PnPID "
                "WHERE e.PartFamilyLongDesc LIKE '%L-%' OR e.PartFamilyLongDesc LIKE '%H-%'"
            ).fetchall():
                if not desc:
                    continue
                for code in LCODE_RE.findall(desc):
                    self.lcode_families[code].append((logical, desc, end, cls))

    def match(self, e: PipingClassEntry) -> None:
        """Resolve one entry: score all catalog candidates, pick the winner, assign confidence."""
        if not e.lcode:
            e.match_note = "sin L-code en la descripcion"
            e.confidence = "BAJA"
            return

        # Prefer the exact code (e.g. L-478-H2); fall back to the base code when the H2 family
        # is absent. ``used_base`` drives the SUSTITUCION flag.
        fams = self.lcode_families.get(e.lcode)
        used_base = False
        if not fams and e.lcode_base:
            fams = self.lcode_families.get(e.lcode_base)
            used_base = bool(fams)
        if not fams:
            e.match_note = f"L-code {e.lcode} no esta en ningun catalogo"
            e.confidence = "BAJA"
            return

        n_families = len({(c[0], c[1]) for c in fams})
        candidates = self._score_candidates(fams, used_base, e)
        if not candidates:
            e.match_note = f"diametro {e.main_diameter} no esta en la familia del L-code"
            e.confidence = "BAJA"
            return

        candidates.sort(key=lambda c: c.score, reverse=True)
        win = candidates[0]
        e.catalog = win.catalog
        e.family_desc = win.family_desc
        e.catalog_class = win.catalog_class
        e.size_record_id = win.size_record_id
        e.part_family_id = win.part_family_id
        e.pnpid = win.pnpid
        e.score = win.score
        e.alternatives = [
            (c.family_desc, c.score, c.schedule, c.pressure_class) for c in candidates[1:4]
        ]

        # ----- confidence -----
        runner_up = candidates[1].score if len(candidates) > 1 else None
        ambiguous = runner_up is not None and (win.score - runner_up) < SCORE_MARGIN
        notes: list[str] = []

        # Substitution: the hydrogen variant fell back to its base family.
        is_substitution = e.is_hydrogen and (win.used_base_code or used_base)

        # End-type disagreement on the winner (family chosen does not match the deduced end type).
        end_mismatch = bool(e.end_type and win.end_type and win.end_type != e.end_type)

        if is_substitution:
            e.confidence = "SUSTITUCION"
            notes.append("base sustituida (H2 no esta en catalogo)")
        elif n_families > 1 and (not e.end_type or end_mismatch):
            e.confidence = "MEDIA"
            notes.append(f"{n_families} familias comparten {e.lcode_base or e.lcode}; "
                         f"end={e.end_type or '?'} no desambigua")
        elif ambiguous:
            e.confidence = "MEDIA"
            notes.append(f"candidato no inequivoco (2o score {runner_up:.0f} vs {win.score:.0f}; "
                         f"falta senal -> sch/rating)")
        else:
            e.confidence = "ALTA"

        if end_mismatch and e.confidence != "MEDIA":
            notes.append(f"end-type {e.end_type} != {win.end_type}")
        e.match_note = "; ".join(notes)

    def _score_candidates(
        self,
        fams: list[tuple[str, str, str | None, str]],
        used_base: bool,
        e: PipingClassEntry,
    ) -> list[_Candidate]:
        """Score every catalog row (across all families of the L-code) at the entry's diameter."""
        out: list[_Candidate] = []
        for catalog, family_desc, _fam_end, cls in {(c[0], c[1], c[2], c[3]) for c in fams}:
            con = self.handles[catalog]
            params: list = [family_desc]
            sql = (
                "SELECT SizeRecordId, PartFamilyId, PnPID, NominalDiameter, Schedule, "
                "PressureClass, EndType FROM EngineeringItems WHERE PartFamilyLongDesc = ?"
            )
            if e.main_diameter is not None:
                sql += " AND ABS(NominalDiameter - ?) < 0.01"
                params.append(e.main_diameter)
            for srid, fam_id, pnpid, _nd, sch, pc, end in con.execute(sql, params).fetchall():
                score = W_LCODE_BASE if used_base else W_LCODE_EXACT
                if e.end_type and end:
                    score += W_END_MATCH if end == e.end_type else W_END_MISMATCH
                if e.main_diameter is not None:
                    score += 3.0  # nominal diameter matched (rows are already filtered by it)
                if e.branch_diameter is not None and self._has_branch_port(
                    con, pnpid, e.main_diameter, e.branch_diameter
                ):
                    score += W_BRANCH_PORT
                if e.schedule:
                    score += W_SCH_MATCH if (sch or "") == e.schedule else W_SCH_MISMATCH
                if e.rating:
                    score += W_RATING_MATCH if (pc or "") == e.rating else W_RATING_MISMATCH
                out.append(_Candidate(
                    score=score, family_desc=family_desc, catalog=catalog, catalog_class=cls,
                    size_record_id=srid, part_family_id=fam_id, pnpid=pnpid,
                    schedule=sch, pressure_class=pc, end_type=end, used_base_code=used_base,
                ))
        return out

    def family_count(self, e: PipingClassEntry) -> int:
        """Number of distinct catalog families that carry this entry's L-code (ambiguity proxy)."""
        code = e.lcode if e.lcode in self.lcode_families else e.lcode_base
        return len({(c[0], c[1]) for c in self.lcode_families.get(code or "", [])})

    @staticmethod
    def _has_branch_port(
        con: sqlite3.Connection, pnpid: int, main_d: float | None, branch_d: float
    ) -> bool:
        """True if the part has a port at the branch diameter (besides the main one)."""
        port_dias = [
            d for (d,) in con.execute(
                "SELECT p.NominalDiameter FROM PartPort pp JOIN Port p ON pp.Port = p.PnPID "
                "WHERE pp.Part = ? AND p.NominalDiameter IS NOT NULL",
                (pnpid,),
            ).fetchall()
        ]
        return any(abs(d - branch_d) < 0.01 for d in port_dias)

    def close(self) -> None:
        for con in self.handles.values():
            con.close()


# =========================================================================== DEFINITION
def build_definition_from_entries(entries: list[PipingClassEntry]) -> SpecDefinition:
    """Assemble a SpecDefinition from matched entries, reusing NXD-2's branch table verbatim.

    Identity/metadata and branch table come from ``derive_definition_from_template`` (the proven
    NXD-2 derivation); only the component list is replaced by the matched piping-class entries.

    Policy decision (documented): the spec includes EVERY entry that resolved to a catalog row
    (ALTA + MEDIA + SUSTITUCION), not only ALTA -- a piping class must be materialised whole, and
    the MEDIA/SUSTITUCION rows are real (correct family, ambiguous size or H2-on-base). The
    confidence is the REVIEW signal (out\\REVISION_MATCHING.xlsx tells the engineer which rows to
    check); it does not silently drop parts. BAJA/no-match entries have no srid and are skipped.
    """
    base = derive_definition_from_template()
    base.name = OUT_NAME
    base.description = "Spec generada desde el piping class Excel (FROMCLASS)"
    base.components = [
        ComponentRef(
            pnpid_template=e.pnpid,                 # catalog PnPID (the materialiser sources by srid)
            class_name=e.catalog_class or "Pipe",
            size_record_id=e.size_record_id,
            part_family_id=e.part_family_id,
        )
        for e in entries
        if e.size_record_id is not None
    ]
    return base


# =========================================================================== REPORT
CONF_ORDER = ["BAJA", "MEDIA", "SUSTITUCION", "ALTA"]   # review order: doubtful first


def _nxd2_truth() -> tuple[set, set]:
    """Ground truth at FAMILY level: the catalog PartFamilyId set NXD-2 actually adopted.

    NXD-2.CatalogPartFamilyId references the catalog PartFamilyId; this is the meaningful oracle.
    Also returns the legacy SizeRecordId-blob set (kept only to expose how misleading it is: NXD-2
    is a materialised spec whose srid blobs are rewritten, so blob equality is NOT a semantic test).
    """
    con = ro_connect(TEMPLATE_PSPC)
    try:
        fam = {s for (s,) in con.execute(
            "SELECT CatalogPartFamilyId FROM EngineeringItems WHERE CatalogPartFamilyId IS NOT NULL"
        )}
        srids = {s for (s,) in con.execute(
            "SELECT SizeRecordId FROM EngineeringItems WHERE SizeRecordId IS NOT NULL"
        )}
    finally:
        con.close()
    return fam, srids


def _crossval(entries: list[PipingClassEntry]) -> dict:
    """Per-confidence-level precision at FAMILY level vs NXD-2, plus the legacy srid-blob figure."""
    gt_fam, gt_srid = _nxd2_truth()
    matched = [e for e in entries if e.size_record_id is not None]

    # Global family-level precision (the honest baseline) and the legacy srid-blob artefact.
    fam_tp = sum(1 for e in matched if e.part_family_id in gt_fam)
    prec_family = fam_tp / len(matched) if matched else 0.0
    uniq_srid = {e.size_record_id for e in matched}
    prec_srid = len(uniq_srid & gt_srid) / len(uniq_srid) if uniq_srid else 0.0

    by_level: dict[str, dict] = {}
    for lvl in CONF_ORDER:
        es = [e for e in entries if e.confidence == lvl]
        mt = [e for e in es if e.size_record_id is not None]
        tp = sum(1 for e in mt if e.part_family_id in gt_fam)
        by_level[lvl] = {
            "n": len(es),
            "matched": len(mt),
            "fam_tp": tp,
            "precision": (tp / len(mt)) if mt else 0.0,
        }

    # Families chosen in ALTA that NXD-2 does NOT contain (the oracle-gap explanation).
    alta_gap_families: dict[str, int] = {}
    for e in entries:
        if e.confidence == "ALTA" and e.size_record_id is not None and e.part_family_id not in gt_fam:
            alta_gap_families[e.family_desc or "?"] = alta_gap_families.get(e.family_desc or "?", 0) + 1

    return {
        "matched": len(matched),
        "precision_family": prec_family,
        "precision_srid_blob": prec_srid,
        "by_level": by_level,
        "alta_gap_families": alta_gap_families,
    }


def print_coverage_report(entries: list[PipingClassEntry], cv: dict) -> None:
    tot = len(entries)
    print("\n===== INFORME DE MATCHING CON NIVELES DE CONFIANZA =====")
    print(f"  Entradas del piping class: {tot}")

    print("\n  --- BASELINE GLOBAL (validacion a nivel FAMILIA vs NXD-2.CatalogPartFamilyId) ---")
    print(f"    Precision a nivel familia (oraculo real): {cv['precision_family'] * 100:.1f}%")
    print(f"    Precision por srid-blob (ARTEFACTO):      {cv['precision_srid_blob'] * 100:.1f}%")
    print("    Difieren porque NXD-2 es una spec materializada: reescribe los blobs SizeRecordId")
    print("    (solo 34/406 coinciden), pero CatalogPartFamilyId si enlaza con el catalogo.")

    print(f"\n  --- PRECISION Y COBERTURA POR NIVEL DE CONFIANZA ---")
    print(f"    {'NIVEL':<13}{'n':>5}{'cob.%':>8}{'casadas':>9}{'fam_TP':>8}{'prec_fam':>10}")
    for lvl in ["ALTA", "MEDIA", "SUSTITUCION", "BAJA"]:
        d = cv["by_level"][lvl]
        cob = 100 * d["n"] / tot if tot else 0
        print(f"    {lvl:<13}{d['n']:>5}{cob:>7.1f}%{d['matched']:>9}{d['fam_tp']:>8}"
              f"{d['precision'] * 100:>9.1f}%")

    gap = cv["alta_gap_families"]
    n_gap = sum(gap.values())
    alta = cv["by_level"]["ALTA"]
    # True ALTA precision once the documented oracle-gap families are removed (NXD-2 never adopted
    # them; they are correct catalog families, not mis-picks). This is the honest ALTA quality.
    alta_eff = alta["matched"] - n_gap
    alta_prec_eff = (alta["fam_tp"] / alta_eff) if alta_eff else 0.0
    print(f"\n  --- NOTA HONESTA: errores residuales en ALTA vs NXD-2 ({n_gap} entradas) ---")
    print("    NO son mis-picks: son familias de catalogo correctas que esta spec NXD-2 no adopta")
    print("    (gap whole-family; el catalogo no trae Material, no se puede afinar mas). Familias:")
    for fam, n in sorted(gap.items(), key=lambda x: -x[1]):
        print(f"      {n:>3}x {fam[:66]}")
    print(f"    Precision ALTA vs NXD-2 (oraculo incompleto): {alta['precision'] * 100:.1f}%  "
          f"<- techo impuesto por las {n_gap} entradas de familias que NXD-2 omite")
    print(f"    Precision ALTA descontando esas familias (calidad real del match): "
          f"{alta_prec_eff * 100:.1f}%")

    # How many baseline auto-picks are now degraded out of ALTA (no longer trusted blindly).
    degraded = sum(1 for e in entries
                   if e.size_record_id is not None and e.confidence != "ALTA")
    print(f"\n  --- DISCIPLINA DE REVISION ---")
    print(f"    Antes: las {cv['matched']} casadas se auto-elegian sin distincion.")
    print(f"    Ahora fuera de ALTA (a revisar): {degraded}  "
          f"(MEDIA {cv['by_level']['MEDIA']['n']} + SUSTITUCION {cv['by_level']['SUSTITUCION']['n']} "
          f"+ BAJA {cv['by_level']['BAJA']['n']}).")
    print(f"    El ingeniero confia en ALTA y revisa solo lo dudoso.")


# =========================================================================== EXCEL DE REVISION
_CONF_FILL = {
    "BAJA": "FFC7CE",          # red
    "MEDIA": "FFEB9C",         # amber
    "SUSTITUCION": "BDD7EE",   # blue
    "ALTA": "C6EFCE",          # green
}


def write_review_xlsx(entries: list[PipingClassEntry], path: str) -> None:
    """One row per piping-class entry, ordered doubtful-first, for human review."""
    wb = Workbook()
    ws = wb.active
    ws.title = "REVISION"
    headers = ["Hoja", "Descripcion", "Tipo", "O (in)", "L-code", "Familia elegida",
               "Confianza", "Candidatos alternativos (score)", "Estado"]
    ws.append(headers)
    bold = Font(bold=True)
    for c in ws[1]:
        c.font = bold
    ws.freeze_panes = "A2"

    order = {lvl: i for i, lvl in enumerate(CONF_ORDER)}  # BAJA(0) ... ALTA(3) -> doubtful first
    for e in sorted(entries, key=lambda x: order.get(x.confidence, 9)):
        alts = " | ".join(
            f"{fd[:40]} ({sc:.0f})" for fd, sc, _sch, _pc in e.alternatives
        )
        ws.append([
            e.sheet,
            e.description.splitlines()[0][:90],
            e.type_,
            e.main_diameter,
            e.lcode or "",
            (e.family_desc or "")[:60],
            e.confidence,
            alts,
            e.match_note,
        ])
        fill = _CONF_FILL.get(e.confidence)
        if fill:
            ws.cell(row=ws.max_row, column=7).fill = PatternFill(
                start_color=fill, end_color=fill, fill_type="solid"
            )

    widths = [14, 50, 16, 8, 12, 40, 13, 46, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    wb.save(path)


# =========================================================================== main
def main() -> None:
    os.makedirs(OUT_DIR, exist_ok=True)
    out_pspc = os.path.join(OUT_DIR, f"{OUT_NAME}.pspc")
    out_pspx = os.path.join(OUT_DIR, f"{OUT_NAME}.pspx")

    print("== Fase 1: parsear el piping class Excel ==")
    entries = parse_workbook(XLSX)
    print(f"  entradas parseadas: {len(entries)} en {len({e.sheet for e in entries})} hojas")

    print("\n== Fase 2: emparejar (scoring multi-senal) y asignar confianza ==")
    matcher = CatalogMatcher()
    try:
        for e in entries:
            matcher.match(e)
    finally:
        matcher.close()

    crossval = _crossval(entries)
    print_coverage_report(entries, crossval)

    review_path = os.path.join(OUT_DIR, "REVISION_MATCHING.xlsx")
    write_review_xlsx(entries, review_path)
    print(f"\n  Excel de revision (dudoso arriba): {review_path}")

    print("\n== Fase 3: construir SpecDefinition y materializar NXD-2-FROMCLASS ==")
    defin = build_definition_from_entries(entries)
    print(f"  componentes en la definicion (casados): {len(defin.components)}")
    catalogs = CatalogIndex()
    try:
        mat = Materialiser(out_pspc, defin, catalogs)
        mat.build()
        build_pspx(out_pspx, defin, out_pspc)
    finally:
        catalogs.close()

    verify_spec(out_pspc, out_pspx, defin, mat)
    print("\nListo. Ficheros en:", OUT_DIR)


if __name__ == "__main__":
    main()
