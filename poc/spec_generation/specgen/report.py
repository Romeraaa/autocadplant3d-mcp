"""The matching review report: ``REVISION_MATCHING.xlsx`` plus a textual coverage summary.

Standalone proof of concept (NOT part of the MCP server). ``openpyxl`` plus :mod:`specgen`.

The report is the human-review signal of the pipeline: one row per piping-class entry, ordered
doubtful-first (BAJA, MEDIA, SUSTITUCION, ALTA), colour-coded by confidence, listing the chosen
catalog family and the runner-up candidates. There is no external oracle (NXD-2 dependency removed):
coverage is reported purely as the distribution of confidence levels and the matched/un-matched
counts, which is what the engineer acts on.
"""

from __future__ import annotations

from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Review order (doubtful first) and per-level fill colours.
CONF_ORDER = ["BAJA", "MEDIA", "SUSTITUCION", "ALTA"]
_CONF_FILL = {
    "BAJA": "FFC7CE",          # red
    "MEDIA": "FFEB9C",         # amber
    "SUSTITUCION": "BDD7EE",   # blue
    "ALTA": "C6EFCE",          # green
}


def coverage(entries) -> dict:
    """Return a structured coverage summary (counts + per-level distribution)."""
    total = len(entries)
    matched = [e for e in entries if e.size_record_id is not None]
    by_level = Counter(e.confidence for e in entries)
    matched_by_level = Counter(e.confidence for e in matched)
    h2 = [e for e in entries if getattr(e, "is_hydrogen", False)]
    h2_dedicated = [e for e in h2 if e.family_desc and "-H2" in e.family_desc]
    # H6: guard the denominator.
    pct = (100.0 * len(matched) / total) if total else 0.0
    return {
        "total": total,
        "matched": len(matched),
        "unmatched": total - len(matched),
        "match_pct": pct,
        "by_level": {lvl: by_level.get(lvl, 0) for lvl in CONF_ORDER},
        "matched_by_level": {lvl: matched_by_level.get(lvl, 0) for lvl in CONF_ORDER},
        "h2_total": len(h2),
        "h2_dedicated_family": len(h2_dedicated),
    }


def format_coverage(cov: dict) -> str:
    """Render the coverage summary as a multi-line string for the CLI."""
    lines = [
        "===== COBERTURA DE MATCHING =====",
        f"  entradas del piping class: {cov['total']}",
        f"  casadas (con SizeRecordId): {cov['matched']} ({cov['match_pct']:.1f}%)  "
        f"sin casar: {cov['unmatched']}",
        "  distribucion por nivel de confianza (total / casadas):",
    ]
    for lvl in CONF_ORDER:
        lines.append(f"    {lvl:<13} {cov['by_level'][lvl]:>5} / "
                     f"{cov['matched_by_level'][lvl]:>5}")
    if cov["h2_total"]:
        lines.append(f"  entradas H2: {cov['h2_total']}  resueltas a familia -H2 dedicada: "
                     f"{cov['h2_dedicated_family']}")
    return "\n".join(lines)


def write_review_xlsx(entries, path: str) -> None:
    """One row per piping-class entry, ordered doubtful-first, for human review."""
    wb = Workbook()
    ws = wb.active
    ws.title = "REVISION"
    headers = ["Hoja", "Descripcion", "Tipo", "O (in)", "L-code", "Familia elegida",
               "Catalogo", "Confianza", "Candidatos alternativos (score)", "Estado"]
    ws.append(headers)
    bold = Font(bold=True)
    for c in ws[1]:
        c.font = bold
    ws.freeze_panes = "A2"

    order = {lvl: i for i, lvl in enumerate(CONF_ORDER)}
    for e in sorted(entries, key=lambda x: order.get(x.confidence, 9)):
        alts = " | ".join(
            f"{(fd or '')[:40]} ({sc:.0f})" for fd, sc, _sch, _pc in e.alternatives
        )
        ws.append([
            e.sheet,
            (e.description.splitlines()[0] if e.description else "")[:90],
            e.type_,
            e.main_diameter,
            e.lcode or "",
            (e.family_desc or "")[:60],
            e.catalog or "",
            e.confidence,
            alts,
            e.match_note,
        ])
        fill = _CONF_FILL.get(e.confidence)
        if fill:
            ws.cell(row=ws.max_row, column=8).fill = PatternFill(
                start_color=fill, end_color=fill, fill_type="solid"
            )

    widths = [14, 50, 16, 8, 12, 40, 26, 13, 46, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    wb.save(path)
